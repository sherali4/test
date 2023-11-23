from datetime import date
import json
import urllib.request
import os
import csv
from config import ids
from openpyxl import Workbook

filename = f"siat-data-{date.today()}.xlsx"

try:
  os.remove(filename)
except OSError:
  print(f"{filename} not exists")

wb = Workbook()
wb.create_sheet("siat_data", 0)
ws = wb.active
birinchi_satr = [
    'id', 'klassifikator_kodi', 'name_uz', 'name_ru', 'name_en', 'davr',
    'miqdor'
]

ws1 = wb['siat_data']
ws = wb.active
row = 1
column = 0
for x in birinchi_satr:
  column = column + 1
  ws1.cell(row=row, column=column, value=x)

for id in ids:
  with urllib.request.urlopen(
      f"https://api.siat.stat.uz/media/uploads/sdmx/sdmx_data_{id}.json"
  ) as url:
    data = json.load(url)
    for d in range(0, len(data[0]['metadata'])):
      for key, value in data[0]['data'][d].items():
        if value == data[0]['data'][d]['Code'] or value == data[0]['data'][d][
            'Klassifikator'] or value == data[0]['data'][d][
                'Klassifikator_ru'] or value == data[0]['data'][d][
                    'Klassifikator_en']:
          pass
        else:
          print(id, data[0]['data'][d]['Code'], ' - ',
                data[0]['data'][d]['Klassifikator'], ' - ',
                data[0]['data'][d]['Klassifikator_ru'], ' - ',
                data[0]['data'][d]['Klassifikator_en'], ' - ', key, ' - ',
                value)
          row = row + 1
          ws1.cell(row=row, column=1, value=id)
          ws1.cell(row=row, column=2, value=data[0]['data'][d]['Code'])
          ws1.cell(row=row,
                   column=3,
                   value=data[0]['data'][d]['Klassifikator'])
          ws1.cell(row=row,
                   column=4,
                   value=data[0]['data'][d]['Klassifikator_ru'])
          ws1.cell(row=row,
                   column=5,
                   value=data[0]['data'][d]['Klassifikator_en'])
          ws1.cell(row=row, column=6, value=key)
          ws1.cell(row=row, column=7, value=value)
wb.save(filename=filename)
if wb.save:
  print("File saved")
wb.close()
