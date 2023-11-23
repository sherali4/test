from datetime import date
import json
import urllib.request
import os
from config import ids

from openpyxl import Workbook

try:
  os.remove(f"siat-{date.today()}.xlsx")
except OSError:
  print(f"siat-{date.today()}.xlsx not exists")
wb = Workbook()
wb.create_sheet("siat", 0)
ws = wb.active
birinchi_satr = [
    'id', 'name_uz', 'name_ru', 'name_en', 'value_uz', 'value_ru', 'value_en'
]

ws1 = wb['siat']
ws = wb.active
row = 1
column = 0
for x in birinchi_satr:
  column = column + 1
  ws1.cell(row=row, column=column, value=x)

for id in ids:
  with urllib.request.urlopen(
      f"https://api.siat.stat.uz/media/uploads/sdmx/sdmx_data_{id}.json"
  ) as url:
    data = json.load(url)
    for d in data[0]['metadata']:
      row = row + 1
      ws1.cell(row=row, column=1, value=id)
      ws1.cell(row=row, column=2, value=d['name_uz'])
      ws1.cell(row=row, column=3, value=d['name_ru'])
      ws1.cell(row=row, column=4, value=d['name_en'])
      ws1.cell(row=row, column=5, value=d['value_uz'])
      ws1.cell(row=row, column=6, value=d['value_ru'])
      ws1.cell(row=row, column=7, value=d['value_en'])

wb.save(f"siat-{date.today()}.xlsx")
if wb.save:
  print("File saved")
wb.close()
