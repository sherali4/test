from datetime import date
import urllib.request
import os
from config import ids, idlar

from openpyxl import Workbook

filename = f"siat-idlar-{date.today()}.xlsx"

try:
  os.remove(filename)
except OSError:
  print(f"{filename} not exists")
wb = Workbook()
wb.create_sheet("siat", 0)
ws = wb.active
birinchi_satr = [
    'id', 'name_uz', 'name_ru', 'name_en', 'value_uz', 'value_ru', 'value_en'
]

ws1 = wb['siat']
ws = wb.active
row = 1
column = 0
for x in birinchi_satr:
  column = column + 1
  ws1.cell(row=row, column=column, value=x)

for id in idlar:
  row = row + 1
  try:
    with urllib.request.urlopen(
        f"https://api.siat.stat.uz/media/uploads/sdmx/sdmx_data_{id}.json"
    ) as url:
      if url:
        ws1.cell(row=row, column=1, value=f'{id}')
        ws1.cell(row=row, column=2, value='bor')
      else:
        ws1.cell(row=row, column=1, value=f'{id}')
        ws1.cell(row=row, column=1, value='yuq')

  except:
    print(f"Error - { id }")

wb.save(filename)
if wb.save:
  print("File saved")
wb.close()
