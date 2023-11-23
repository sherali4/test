from datetime import date
import json
import urllib.request
import os
import csv

from openpyxl import Workbook

filename = f"siat-data-{date.today()}.xlsx"
csv_filename = f"siat-data-{date.today()}.csv"

try:
  os.remove(csv_filename)
except OSError:
  print(f"{csv_filename} not exists")

# wb = Workbook()
# wb.create_sheet("siat_data", 0)
# ws = wb.active
birinchi_satr = [
    'id', 'klassifikator_kodi', 'name_uz', 'name_ru', 'name_en', 'davr',
    'miqdor'
]

# ws1 = wb['siat_data']
# ws = wb.active
# row = 1
# column = 0
# for x in birinchi_satr:
#   column = column + 1
#   ws1.cell(row=row, column=column, value=x)

ids = [
    '420', '456', '374', '461', '388', '435', '357', '1597', '1596', '454',
    '361', '466', '396', '491', '416', '476', '411', '433', '303', '296',
    '469', '402', '430', '276', '274', '425', '2693', '1601', '3073', '1602',
    '3074', '1606', '1608', '1611', '1620', '1631', '1646', '1655', '1660',
    '1665', '1670', '1672', '3359', '3358', '3356', '3357', '3355'
]

with open(csv_filename, 'w', encoding='utf-8', newline='') as f:
  writer = csv.writer(f, delimiter=';')
  writer.writerow(birinchi_satr)
  for id in ids:
    url = f"https://api.siat.stat.uz/media/uploads/sdmx/sdmx_data_{id}.json"
    # with urllib.request.urlopen(
    #     f"https://api.siat.stat.uz/media/uploads/sdmx/sdmx_data_{id}.json"
    # ) as url:
    #   data = json.load(url)

    with urllib.request.urlopen(url) as response:
      data = json.loads(response.read().decode('utf-8'))

      for d in range(0, len(data[0]['metadata'])):
        for key, value in data[0]['data'][d].items():
          if value == data[0]['data'][d]['Code'] or value == data[0]['data'][
              d]['Klassifikator'] or value == data[0]['data'][d][
                  'Klassifikator_ru'] or value == data[0]['data'][d][
                      'Klassifikator_en']:
            pass
          else:
            writer.writerow([
                id,
                data[0]['data'][d]['Code'],
                data[0]['data'][d]['Klassifikator'],
                data[0]['data'][d]['Klassifikator_ru'],
                data[0]['data'][d]['Klassifikator_en'],
                key,
                value,
            ])

print('successfully')

# for id in ids:
#   with urllib.request.urlopen(
#       f"https://api.siat.stat.uz/media/uploads/sdmx/sdmx_data_{id}.json"
#   ) as url:
#     data = json.load(url)
#     for d in range(0, len(data[0]['metadata'])):
#       for key, value in data[0]['data'][d].items():
#         if value == data[0]['data'][d]['Code'] or value == data[0]['data'][d][
#             'Klassifikator'] or value == data[0]['data'][d][
#                 'Klassifikator_ru'] or value == data[0]['data'][d][
#                     'Klassifikator_en']:
#           pass
#         else:
#           print(id, data[0]['data'][d]['Code'], ' - ',
#                 data[0]['data'][d]['Klassifikator'], ' - ',
#                 data[0]['data'][d]['Klassifikator_ru'], ' - ',
#                 data[0]['data'][d]['Klassifikator_en'], ' - ', key, ' - ',
#                 value)
#           row = row + 1
#           ws1.cell(row=row, column=1, value=id)
#           ws1.cell(row=row, column=2, value=data[0]['data'][d]['Code'])
#           ws1.cell(row=row,
#                    column=3,
#                    value=data[0]['data'][d]['Klassifikator'])
#           ws1.cell(row=row,
#                    column=4,
#                    value=data[0]['data'][d]['Klassifikator_ru'])
#           ws1.cell(row=row,
#                    column=5,
#                    value=data[0]['data'][d]['Klassifikator_en'])
#           ws1.cell(row=row, column=6, value=key)
#           ws1.cell(row=row, column=7, value=value)
# wb.save(filename=filename)
# if wb.save:
#   print("File saved")
# wb.close()
